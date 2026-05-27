"""Tests pra comparisons + workloads + exporters (Chunk 1.3)."""

import io

import pytest
from openpyxl import load_workbook

from data_agents.cost_app.databricks.comparisons import (
    compute_comparison,
    get_summary_table,
)
from data_agents.cost_app.databricks.exporters import (
    build_xlsx_multi_scenarios,
    build_xlsx_single_scenario,
    suggest_filename,
)
from data_agents.cost_app.databricks.workloads import (
    aggregate_workloads,
)
from data_agents.cost_app.databricks.workloads import (
    get_summary_table as workloads_summary,
)
from data_agents.cost_engine.databricks import DatabricksScenario


# ─── Fixtures ────────────────────────────────────────────────────────────────


@pytest.fixture
def small_scenario() -> DatabricksScenario:
    return DatabricksScenario(
        cloud="azure",
        compute_type="jobs_compute",
        tier="standard",
        photon=False,
        driver_instance="Standard_DS3_v2",
        worker_instance="Standard_DS3_v2",
        num_workers=1,
        hours_per_day=1,
        days_per_month=22,
        region="brazilsouth",
        instance_pricing_model="on_demand",
        driver_instance_cost_per_hour_usd=0.286,
        worker_instance_cost_per_hour_usd=0.286,
    )


@pytest.fixture
def large_scenario() -> DatabricksScenario:
    return DatabricksScenario(
        cloud="azure",
        compute_type="jobs_compute",
        tier="premium",
        photon=True,
        driver_instance="Standard_E16ds_v4",
        worker_instance="Standard_E16ds_v4",
        num_workers=10,
        hours_per_day=24,
        days_per_month=30,
        region="brazilsouth",
        instance_pricing_model="on_demand",
        driver_instance_cost_per_hour_usd=1.260,
        worker_instance_cost_per_hour_usd=1.260,
    )


@pytest.fixture
def aws_scenario() -> DatabricksScenario:
    return DatabricksScenario(
        cloud="aws",
        compute_type="all_purpose_compute",
        tier="premium",
        photon=False,
        driver_instance="m5.xlarge",
        worker_instance="m5.xlarge",
        num_workers=2,
        hours_per_day=8,
        days_per_month=22,
        region="us-east-1",
        instance_pricing_model="on_demand",
        driver_instance_cost_per_hour_usd=0.192,
        worker_instance_cost_per_hour_usd=0.192,
    )


# ─── Comparisons ─────────────────────────────────────────────────────────────


class TestComparisons:
    def test_small_workload_no_savings(self, small_scenario):
        """Workload pequeno (<$10k/ano DBU) → sem DBCU discount."""
        result = compute_comparison(small_scenario)
        assert result.savings_1y_annual == 0
        assert result.savings_3y_annual == 0
        assert result.savings_1y_pct == 0
        assert result.savings_3y_pct == 0
        assert "Permaneça em Pay-as-you-go" in result.recommendation
        assert result.breakeven_month_1y is None
        assert result.breakeven_month_3y is None

    def test_large_workload_has_savings(self, large_scenario):
        result = compute_comparison(large_scenario)
        assert result.savings_1y_annual > 0
        assert result.savings_3y_annual > result.savings_1y_annual
        assert result.savings_3y_pct > 0
        # Recomendação deve mencionar DBCU
        assert "DBCU" in result.recommendation

    def test_cumulative_36m_has_36_entries(self, small_scenario):
        result = compute_comparison(small_scenario)
        assert len(result.cumulative_36m) == 36
        # Month 1 deve ser igual a monthly_payg
        assert result.cumulative_36m[0]["payg"] == pytest.approx(result.monthly_payg)
        # Month 36 deve ser approx 36 × monthly
        assert result.cumulative_36m[35]["payg"] == pytest.approx(
            result.monthly_payg * 36, rel=0.01
        )

    def test_breakeven_when_savings_exist(self, large_scenario):
        result = compute_comparison(large_scenario)
        # Pra workload grande com DBCU benefits, breakeven deve aparecer cedo
        # (no extremo, mês 1 já tem cumulative menor com DBCU)
        if result.savings_1y_pct > 0:
            assert result.breakeven_month_1y is not None
            assert 1 <= result.breakeven_month_1y <= 36
        if result.savings_3y_pct > 0:
            assert result.breakeven_month_3y is not None

    def test_summary_table_has_3_options(self, large_scenario):
        result = compute_comparison(large_scenario)
        table = get_summary_table(result)
        assert len(table) == 3
        opcoes = [row["Opção"] for row in table]
        assert opcoes == ["Pay-as-you-go", "DBCU Commit 1 ano", "DBCU Commit 3 anos"]

    def test_currency_propagates(self, small_scenario):
        small_scenario.currency_label = "BRL"
        small_scenario.currency_conversion_rate = 5.0
        result = compute_comparison(small_scenario)
        assert result.currency == "BRL"


# ─── Workloads aggregation ──────────────────────────────────────────────────


class TestWorkloadsAggregation:
    def test_aggregate_two_workloads_sums_monthly(self, small_scenario, large_scenario):
        workloads = [
            ("ETL Bronze", "Pipeline noturno", small_scenario),
            ("ETL Silver", "Pipeline grande 24/7", large_scenario),
        ]
        agg = aggregate_workloads(workloads)
        assert len(agg.workloads) == 2
        assert agg.total_monthly > 0
        # Soma deve ser igual ao individual de cada
        from data_agents.cost_engine.databricks import calculate_databricks_cost

        m1 = calculate_databricks_cost(small_scenario)["totals"]["monthly"]
        m2 = calculate_databricks_cost(large_scenario)["totals"]["monthly"]
        assert agg.total_monthly == pytest.approx(m1 + m2, rel=0.001)

    def test_aggregate_empty_raises(self):
        with pytest.raises(ValueError, match="vazia"):
            aggregate_workloads([])

    def test_aggregate_mixed_currencies_raises(self, small_scenario):
        s_brl = DatabricksScenario(
            **{
                **small_scenario.__dict__,
                "currency_label": "BRL",
                "currency_conversion_rate": 5.0,
            }
        )
        with pytest.raises(ValueError, match="currencies diferentes"):
            aggregate_workloads([("A", "", small_scenario), ("B", "", s_brl)])

    def test_aggregate_by_compute_type(self, small_scenario, aws_scenario):
        # small_scenario = jobs_compute / aws_scenario = all_purpose_compute
        workloads = [
            ("Jobs1", "", small_scenario),
            ("Jobs2", "", small_scenario),
            ("Interactive", "", aws_scenario),
        ]
        agg = aggregate_workloads(workloads)
        assert "jobs_compute" in agg.by_compute_type
        assert "all_purpose_compute" in agg.by_compute_type
        # jobs_compute deve ter contribuição de 2 workloads
        from data_agents.cost_engine.databricks import calculate_databricks_cost

        m_jobs = calculate_databricks_cost(small_scenario)["totals"]["monthly"]
        assert agg.by_compute_type["jobs_compute"] == pytest.approx(m_jobs * 2, rel=0.001)

    def test_aggregate_by_cloud(self, small_scenario, aws_scenario):
        # small = azure, aws_scenario = aws
        workloads = [
            ("Azure1", "", small_scenario),
            ("AWS1", "", aws_scenario),
        ]
        agg = aggregate_workloads(workloads)
        assert "azure" in agg.by_cloud
        assert "aws" in agg.by_cloud

    def test_summary_table_includes_total_row(self, small_scenario, large_scenario):
        workloads = [("A", "", small_scenario), ("B", "", large_scenario)]
        agg = aggregate_workloads(workloads)
        table = workloads_summary(agg)
        # 2 workloads + 1 row TOTAL = 3
        assert len(table) == 3
        assert table[-1]["Workload"] == "**TOTAL**"
        assert table[-1]["% do Total"] == 100.0

    def test_failed_workload_emits_warning(self):
        # Cria um scenario inválido
        bad = DatabricksScenario(
            cloud="azure",
            compute_type="jobs_compute",
            tier="premium",
            photon=False,
            driver_instance="NonExistent_SKU",
            worker_instance="Standard_DS3_v2",
            num_workers=1,
            hours_per_day=1,
            days_per_month=1,
            region="brazilsouth",
            instance_pricing_model="on_demand",
            driver_instance_cost_per_hour_usd=0.10,
            worker_instance_cost_per_hour_usd=0.10,
        )
        agg = aggregate_workloads([("Bad", "Invalid SKU", bad)])
        assert len(agg.warnings) == 1
        assert "Bad" in agg.warnings[0]
        assert agg.total_monthly == 0  # nenhum válido


# ─── Exporters XLSX ──────────────────────────────────────────────────────────


class TestExportersXLSX:
    def test_single_scenario_returns_bytesio(self, small_scenario):
        output = build_xlsx_single_scenario("Test Scenario", small_scenario)
        assert isinstance(output, io.BytesIO)
        # Deve ter conteúdo > 0
        output.seek(0, 2)  # end
        assert output.tell() > 1000  # XLSX vazio tem ~5KB; com conteúdo > 1KB
        output.seek(0)

    def test_xlsx_has_required_sheets(self, small_scenario):
        output = build_xlsx_single_scenario("Test", small_scenario)
        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        assert "Resumo Executivo" in sheet_names
        assert "Cenários Detalhados" in sheet_names
        assert "DBCU Comparison" in sheet_names
        assert "Breakdown Hourly" in sheet_names

    def test_xlsx_resumo_has_total_row(self, small_scenario, large_scenario):
        output = build_xlsx_multi_scenarios(
            [("A", small_scenario), ("B", large_scenario)]
        )
        wb = load_workbook(output)
        ws = wb["Resumo Executivo"]
        # Procura linha com "TOTAL" na coluna B
        found_total = False
        for row in ws.iter_rows(values_only=True):
            if row and "TOTAL" in str(row):
                found_total = True
                break
        assert found_total

    def test_xlsx_with_aggregate_adds_extra_sheet(self, small_scenario, aws_scenario):
        workloads = [
            ("A", "", small_scenario),
            ("B", "", aws_scenario),
        ]
        # AWS scenario tem currency USD igual small → ok pra aggregate
        agg = aggregate_workloads(workloads)
        output = build_xlsx_multi_scenarios(
            [("A", small_scenario), ("B", aws_scenario)],
            aggregate=agg,
        )
        wb = load_workbook(output)
        assert "Workload Aggregate" in wb.sheetnames

    def test_xlsx_empty_scenarios_raises(self):
        with pytest.raises(ValueError, match="vazia"):
            build_xlsx_multi_scenarios([])

    def test_xlsx_mixed_currencies_raises(self, small_scenario):
        s_brl = DatabricksScenario(
            **{
                **small_scenario.__dict__,
                "currency_label": "BRL",
                "currency_conversion_rate": 5.0,
            }
        )
        with pytest.raises(ValueError, match="currencies diferentes"):
            build_xlsx_multi_scenarios([("A", small_scenario), ("B", s_brl)])

    def test_suggest_filename_has_timestamp(self):
        fname = suggest_filename()
        assert fname.startswith("databricks-cost-")
        assert fname.endswith(".xlsx")
        # Format: databricks-cost-YYYYMMDD-HHMM.xlsx
        assert len(fname) == len("databricks-cost-20260526-1234.xlsx")

    def test_suggest_filename_custom_prefix(self):
        fname = suggest_filename(prefix="my-quote")
        assert fname.startswith("my-quote-")

    def test_xlsx_detalhes_has_all_columns(self, small_scenario):
        output = build_xlsx_single_scenario("Test", small_scenario)
        wb = load_workbook(output)
        ws = wb["Cenários Detalhados"]
        # Header row deve ter 20 colunas (vide implementação)
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 20
        assert "Cloud" in headers
        assert "Photon" in headers
        assert "Pricing Model" in headers

    def test_xlsx_brl_currency_format(self, small_scenario):
        small_scenario.currency_label = "BRL"
        small_scenario.currency_conversion_rate = 5.0
        output = build_xlsx_single_scenario("Test BRL", small_scenario)
        wb = load_workbook(output)
        ws = wb["Resumo Executivo"]
        # Procura a linha com valores numéricos e checa format
        found_brl_format = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.number_format and "R$" in cell.number_format:
                    found_brl_format = True
                    break
            if found_brl_format:
                break
        assert found_brl_format
