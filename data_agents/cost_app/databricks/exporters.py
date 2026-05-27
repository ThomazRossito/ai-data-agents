"""
Exporters — XLSX para Databricks Cost Calculator.

Gera planilhas formatadas via openpyxl. Output em memória (BytesIO) pra
download via st.download_button.

Sheets gerados:
  1. Resumo Executivo — visão executiva pra cliente
  2. Cenários Detalhados — 1 row por scenario com todos os campos
  3. DBCU Comparison — PAYG vs 1y vs 3y com breakeven
  4. Breakdown Hourly — DBU vs Instance por scenario
  5. Workload Aggregate (opcional, só se houver múltiplos)
"""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_agents.cost_app.databricks.comparisons import compute_comparison
from data_agents.cost_app.databricks.workloads import WorkloadAggregate
from data_agents.cost_engine.databricks import (
    DatabricksScenario,
    calculate_databricks_cost,
)


# ─── Styling helpers ─────────────────────────────────────────────────────────


_HEADER_FILL = PatternFill(start_color="1A1F2C", end_color="1A1F2C", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
_SUBHEADER_FONT = Font(bold=True, color="0E1117", size=10)
_TOTAL_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
_CURRENCY_FORMAT = '_-"$" * #,##0.00_-;-"$" * #,##0.00_-;_-"$" * "-"??_-;_-@_-'
_CURRENCY_FORMAT_BRL = '_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
_PCT_FORMAT = "0.0%"


def _apply_header(ws, row: int, ncols: int) -> None:
    """Aplica style de header (bold + fill escuro) a uma row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_total(ws, row: int, ncols: int) -> None:
    """Aplica style de total row (fill vermelho coral)."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT


def _auto_width(ws, max_width: int = 35) -> None:
    """Ajusta width das colunas baseado no conteúdo (capa em max_width)."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                pass
        adjusted = min(max_length + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _money_format(currency: str) -> str:
    return _CURRENCY_FORMAT_BRL if currency == "BRL" else _CURRENCY_FORMAT


# ─── Sheets ──────────────────────────────────────────────────────────────────


def _write_sheet_resumo(
    wb: Workbook,
    scenarios: list[tuple[str, DatabricksScenario]],
    currency: str,
) -> None:
    """Sheet 1: Resumo Executivo — visão de 1 página pra cliente."""
    ws = wb.create_sheet("Resumo Executivo", 0)
    money_fmt = _money_format(currency)

    # Header titulo
    ws.merge_cells("A1:G1")
    ws["A1"] = "💰 Databricks Cost — Resumo Executivo"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Metadata
    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A4"] = "Cenários:"
    ws["B4"] = len(scenarios)
    ws["A5"] = "Currency:"
    ws["B5"] = currency

    # Header da tabela
    row = 7
    headers = ["#", "Nome", "Cloud", "Compute", "Mensal", "Anual", "TCO 36m"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _apply_header(ws, row, len(headers))

    # Rows
    total_monthly = 0.0
    total_annual = 0.0
    for idx, (name, scenario) in enumerate(scenarios, start=1):
        row += 1
        result = calculate_databricks_cost(scenario)
        m = result["totals"]["monthly"]
        a = result["totals"]["annual"]
        tco = result["totals"]["tco_36m"]
        total_monthly += m
        total_annual += a

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=scenario.cloud.upper())
        ws.cell(row=row, column=4, value=scenario.compute_type)
        for col in range(5, 8):
            ws.cell(row=row, column=col).number_format = money_fmt
        ws.cell(row=row, column=5, value=m)
        ws.cell(row=row, column=6, value=a)
        ws.cell(row=row, column=7, value=tco)

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="TOTAL")
    ws.cell(row=row, column=3, value="—")
    ws.cell(row=row, column=4, value=f"{len(scenarios)} workload(s)")
    ws.cell(row=row, column=5, value=total_monthly).number_format = money_fmt
    ws.cell(row=row, column=6, value=total_annual).number_format = money_fmt
    ws.cell(row=row, column=7, value=total_annual * 3).number_format = money_fmt
    _apply_total(ws, row, len(headers))

    _auto_width(ws)


def _write_sheet_cenarios_detalhados(
    wb: Workbook,
    scenarios: list[tuple[str, DatabricksScenario]],
    currency: str,
) -> None:
    """Sheet 2: cada cenário com TODOS os campos do DatabricksScenario."""
    ws = wb.create_sheet("Cenários Detalhados")
    money_fmt = _money_format(currency)

    # Header
    headers = [
        "Nome",
        "Cloud",
        "Compute Type",
        "Tier",
        "Photon",
        "Driver Instance",
        "Worker Instance",
        "Num Workers",
        "Hours/Day",
        "Days/Month",
        "Region",
        "Pricing Model",
        "Driver $/h",
        "Worker $/h",
        "Autoscale %",
        "DBU/h",
        "Cluster Total/h (USD)",
        "Monthly",
        "Annual",
        "TCO 36m",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _apply_header(ws, 1, len(headers))

    for idx, (name, scenario) in enumerate(scenarios, start=2):
        result = calculate_databricks_cost(scenario)
        breakdown = result["breakdown_hourly_usd"]

        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=scenario.cloud.upper())
        ws.cell(row=idx, column=3, value=scenario.compute_type)
        ws.cell(row=idx, column=4, value=scenario.tier)
        ws.cell(row=idx, column=5, value="ON" if scenario.photon else "OFF")
        ws.cell(row=idx, column=6, value=scenario.driver_instance)
        ws.cell(row=idx, column=7, value=scenario.worker_instance)
        ws.cell(row=idx, column=8, value=scenario.num_workers)
        ws.cell(row=idx, column=9, value=scenario.hours_per_day)
        ws.cell(row=idx, column=10, value=scenario.days_per_month)
        ws.cell(row=idx, column=11, value=scenario.region)
        ws.cell(row=idx, column=12, value=scenario.instance_pricing_model)
        ws.cell(
            row=idx, column=13, value=scenario.driver_instance_cost_per_hour_usd
        ).number_format = _CURRENCY_FORMAT
        ws.cell(
            row=idx, column=14, value=scenario.worker_instance_cost_per_hour_usd
        ).number_format = _CURRENCY_FORMAT
        ws.cell(row=idx, column=15, value=scenario.autoscale_avg_workers_pct)
        ws.cell(row=idx, column=16, value=round(breakdown["dbu_total"], 4))
        ws.cell(
            row=idx, column=17, value=round(breakdown["cluster_total"], 4)
        ).number_format = _CURRENCY_FORMAT
        ws.cell(row=idx, column=18, value=result["totals"]["monthly"]).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=19, value=result["totals"]["annual"]).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=20, value=result["totals"]["tco_36m"]).number_format = (
            money_fmt
        )

    _auto_width(ws, max_width=22)


def _write_sheet_dbcu_comparison(
    wb: Workbook,
    scenarios: list[tuple[str, DatabricksScenario]],
    currency: str,
) -> None:
    """Sheet 3: PAYG vs DBCU 1y vs 3y por scenario, com savings."""
    ws = wb.create_sheet("DBCU Comparison")
    money_fmt = _money_format(currency)

    headers = [
        "Nome",
        "Monthly PAYG",
        "Monthly DBCU 1y",
        "Monthly DBCU 3y",
        "Savings 1y/ano",
        "Savings 3y/ano",
        "Savings 1y %",
        "Savings 3y %",
        "Breakeven 1y (mês)",
        "Breakeven 3y (mês)",
        "Recomendação",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _apply_header(ws, 1, len(headers))

    for idx, (name, scenario) in enumerate(scenarios, start=2):
        comparison = compute_comparison(scenario)

        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=comparison.monthly_payg).number_format = money_fmt
        ws.cell(row=idx, column=3, value=comparison.monthly_dbcu_1y).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=4, value=comparison.monthly_dbcu_3y).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=5, value=comparison.savings_1y_annual).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=6, value=comparison.savings_3y_annual).number_format = (
            money_fmt
        )
        ws.cell(row=idx, column=7, value=comparison.savings_1y_pct / 100).number_format = (
            _PCT_FORMAT
        )
        ws.cell(row=idx, column=8, value=comparison.savings_3y_pct / 100).number_format = (
            _PCT_FORMAT
        )
        ws.cell(
            row=idx, column=9, value=comparison.breakeven_month_1y or "nunca"
        )
        ws.cell(
            row=idx, column=10, value=comparison.breakeven_month_3y or "nunca"
        )
        ws.cell(row=idx, column=11, value=comparison.recommendation)

    _auto_width(ws, max_width=50)


def _write_sheet_breakdown(
    wb: Workbook,
    scenarios: list[tuple[str, DatabricksScenario]],
) -> None:
    """Sheet 4: breakdown hourly DBU vs Instance por scenario."""
    ws = wb.create_sheet("Breakdown Hourly")

    headers = [
        "Nome",
        "Cloud",
        "DBU Driver/h",
        "DBU Workers/h",
        "DBU Total/h",
        "Instance Driver/h",
        "Instance Workers/h",
        "Instance Total/h",
        "Cluster Total/h",
        "% DBU",
        "% Instance",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _apply_header(ws, 1, len(headers))

    for idx, (name, scenario) in enumerate(scenarios, start=2):
        result = calculate_databricks_cost(scenario)
        b = result["breakdown_hourly_usd"]
        total = b["cluster_total"]
        dbu_pct = b["dbu_total"] / total if total > 0 else 0.0
        inst_pct = b["instance_total"] / total if total > 0 else 0.0

        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=scenario.cloud.upper())
        for col, val in [
            (3, b["dbu_driver"]),
            (4, b["dbu_workers"]),
            (5, b["dbu_total"]),
            (6, b["instance_driver"]),
            (7, b["instance_workers"]),
            (8, b["instance_total"]),
            (9, b["cluster_total"]),
        ]:
            ws.cell(row=idx, column=col, value=round(val, 4)).number_format = (
                _CURRENCY_FORMAT
            )
        ws.cell(row=idx, column=10, value=dbu_pct).number_format = _PCT_FORMAT
        ws.cell(row=idx, column=11, value=inst_pct).number_format = _PCT_FORMAT

    _auto_width(ws)


def _write_sheet_aggregate(wb: Workbook, aggregate: WorkloadAggregate) -> None:
    """Sheet opcional: breakdown agregado por compute_type + cloud."""
    ws = wb.create_sheet("Workload Aggregate")
    money_fmt = _money_format(aggregate.currency)

    # Section 1: por compute_type
    ws["A1"] = "Por Compute Type"
    ws["A1"].font = _SUBHEADER_FONT
    ws["A1"].fill = _SUBHEADER_FILL

    ws["A3"] = "Compute Type"
    ws["B3"] = "Total Mensal"
    ws["C3"] = "% do Total"
    _apply_header(ws, 3, 3)

    row = 4
    total = aggregate.total_monthly
    for ct, value in sorted(
        aggregate.by_compute_type.items(), key=lambda x: -x[1]
    ):
        ws.cell(row=row, column=1, value=ct)
        ws.cell(row=row, column=2, value=value).number_format = money_fmt
        pct = value / total if total > 0 else 0.0
        ws.cell(row=row, column=3, value=pct).number_format = _PCT_FORMAT
        row += 1

    # Section 2: por cloud
    row += 2
    ws.cell(row=row, column=1, value="Por Cloud").font = _SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = _SUBHEADER_FILL

    row += 2
    ws.cell(row=row, column=1, value="Cloud")
    ws.cell(row=row, column=2, value="Total Mensal")
    ws.cell(row=row, column=3, value="% do Total")
    _apply_header(ws, row, 3)

    for cl, value in sorted(aggregate.by_cloud.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=cl.upper())
        ws.cell(row=row, column=2, value=value).number_format = money_fmt
        pct = value / total if total > 0 else 0.0
        ws.cell(row=row, column=3, value=pct).number_format = _PCT_FORMAT

    # Section 3: DBU vs Instance global
    row += 3
    ws.cell(row=row, column=1, value="DBU vs Instance (mensal)").font = _SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = _SUBHEADER_FILL

    row += 2
    ws.cell(row=row, column=1, value="Categoria")
    ws.cell(row=row, column=2, value="Total Mensal")
    _apply_header(ws, row, 2)

    row += 1
    ws.cell(row=row, column=1, value="DBU Cost")
    ws.cell(row=row, column=2, value=aggregate.dbu_total_monthly).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value="Instance Cost")
    ws.cell(
        row=row, column=2, value=aggregate.instance_total_monthly
    ).number_format = money_fmt

    _auto_width(ws)


# ─── Public API ──────────────────────────────────────────────────────────────


def build_xlsx_single_scenario(
    name: str,
    scenario: DatabricksScenario,
) -> BytesIO:
    """Build XLSX pra 1 scenario só. Sheets: Resumo + Detalhes + DBCU + Breakdown."""
    return build_xlsx_multi_scenarios([(name, scenario)])


def build_xlsx_multi_scenarios(
    scenarios: list[tuple[str, DatabricksScenario]],
    aggregate: WorkloadAggregate | None = None,
) -> BytesIO:
    """
    Build XLSX pra N scenarios.

    Args:
        scenarios: lista de (name, scenario).
        aggregate: opcional, se passar adiciona Sheet 5 de agregação.

    Returns:
        BytesIO com o XLSX serializado, pronto pra st.download_button.

    Raises:
        ValueError: se scenarios vazio ou misturar currencies diferentes.
    """
    if not scenarios:
        raise ValueError("Lista de scenarios vazia.")

    currencies = {s.currency_label for _, s in scenarios}
    if len(currencies) > 1:
        raise ValueError(
            f"Scenarios usam currencies diferentes ({currencies}). "
            "Use a mesma currency."
        )
    currency = scenarios[0][1].currency_label

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_sheet_resumo(wb, scenarios, currency)
    _write_sheet_cenarios_detalhados(wb, scenarios, currency)
    _write_sheet_dbcu_comparison(wb, scenarios, currency)
    _write_sheet_breakdown(wb, scenarios)
    if aggregate is not None:
        _write_sheet_aggregate(wb, aggregate)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def suggest_filename(prefix: str = "databricks-cost") -> str:
    """Sugere filename com timestamp UTC."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    return f"{prefix}-{ts}.xlsx"
